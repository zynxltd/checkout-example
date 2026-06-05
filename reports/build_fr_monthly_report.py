#!/usr/bin/env python3
"""Generate Fresh Relevance monthly performance report template."""

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPORT_DIR = Path(__file__).resolve().parent
OUTPUT = REPORT_DIR / "Fresh-Relevance-Monthly-Report.xlsx"
SLOTS_CSV = REPORT_DIR / "data" / "fr-slots-2025-06-to-2026-06.csv"
SLOTS_PERIOD = "1 Jun 2025 – 1 Jun 2026"

MTD_YEAR = 2026
MTD_MONTH = "Jun"
MTD_FILL = "FFFFF2CC"
HEAT_UP_FILL = "FFE2F0D9"
DATA_FONT = Font(color="000000")
DATA_FONT_BOLD = Font(color="000000", bold=True)
HEAT_DOWN_FILL = "FFFCE4D6"
MISSING_FILL = "FFFFC7CE"  # faded red — data not available from FR screenshots
CLICK_TO_ORDER = 0.03

# Each entry: (year, month, data dict or None for placeholder row)
REPORT_ROWS = [
    # 2025 — Jun–Aug from FR dashboard screenshots
    (2025, "Jun", {
        "site": 254_360.01, "web": 4_544.63, "email": 0,
        "web_impressions": 21_540, "email_impressions": 0,
        "total_aov": 31.50, "web_aov": 55.42, "email_aov": 0,
        "total_conversion": 0.03, "web_conversion": 0.0038, "email_conversion": 0,
        "notes": "Email FR blocks not active; web-only month.",
    }),
    (2025, "Jul", {
        "site": 720_264.93, "web": 64_405.92, "email": 17_580.83,
        "web_impressions": 1_245_860, "email_impressions": 3_799_747,
        "total_aov": 33.64, "web_aov": 37.01, "email_aov": 30.68,
        "total_conversion": 0.0246, "web_conversion": 0.0014, "email_conversion": 0.0002,
        "notes": "Web revenue scales up; email FR blocks live.",
    }),
    (2025, "Aug", {
        "site": 816_098.55, "web": 18_322.53, "email": 26_699.06,
        "web_impressions": 1_391_355, "email_impressions": 7_185_406,
        "total_aov": 33.58, "web_aov": 23.46, "email_aov": 34.01,
        "total_conversion": 0.0267, "web_conversion": 0.0006, "email_conversion": 0.0001,
        "notes": "Email overtakes web in attributed revenue.",
    }),
    (2025, "Sep", {
        "site": 630_881.06, "web": 14_992.44, "email": 17_574.32,
        "web_impressions": 1_554_518, "email_impressions": 6_593_412,
        "total_aov": 33.67, "web_aov": 24.82, "email_aov": 33.99,
        "total_conversion": 0.0266, "web_conversion": 0.0004, "email_conversion": 0.0001,
        "notes": "Revenue down vs Aug; web impressions up.",
    }),
    (2025, "Oct", {
        "site": 473_355.22, "web": 11_249.06, "email": 10_395.11,
        "web_impressions": 1_181_665, "email_impressions": 5_901_396,
        "total_aov": 32.26, "web_aov": 21.97, "email_aov": 34.31,
        "total_conversion": 0.0222, "web_conversion": 0.0004, "email_conversion": 0.0001,
        "notes": "Down vs prior months across channels.",
    }),
    (2025, "Nov", {
        "site": 347_624.65, "web": 9_402.59, "email": 8_413.31,
        "web_impressions": 796_807, "email_impressions": 7_515_711,
        "total_aov": 32.25, "web_aov": 24.94, "email_aov": 35.20,
        "total_conversion": 0.0224, "web_conversion": 0.0005, "email_conversion": 0,
        "notes": "Email impressions up; revenue down vs Oct.",
    }),
    (2025, "Dec", {
        "site": 363_287.36, "web": 7_062.99, "email": 6_182.28,
        "web_impressions": 586_710, "email_impressions": 5_207_030,
        "total_aov": 23.20, "web_aov": 19.84, "email_aov": 30.01,
        "total_conversion": 0.0429, "web_conversion": 0.0006, "email_conversion": 0,
        "notes": "Lowest AOV month in H2 2025.",
    }),
    # 2026 — Jan–Jun from FR dashboard
    (2026, "Jan", {
        "site": 426_480.98, "web": 13_343.22, "email": 12_034.64,
        "web_impressions": 1_756_199, "email_impressions": 6_077_003,
        "total_aov": 33.52, "web_aov": 30.33, "email_aov": 34.19,
        "total_conversion": 0.0256, "web_conversion": 0.0003, "email_conversion": 0.0001,
        "notes": "Web + email baseline month.",
    }),
    (2026, "Feb", {
        "site": 702_685.35, "web": 11_842.22, "email": 24_657.79,
        "web_impressions": 2_364_555, "email_impressions": 6_648_876,
        "total_aov": 34.30, "web_aov": 24.12, "email_aov": 33.28,
        "total_conversion": 0.0364, "web_conversion": 0.0002, "email_conversion": 0.0001,
        "notes": "Email revenue up; web revenue dipped vs Jan.",
    }),
    (2026, "Mar", {
        "site": 1_505_634.59, "web": 56_827.99, "email": 42_045.60,
        "web_impressions": 6_553_245, "email_impressions": 8_129_826,
        "total_aov": 34.67, "web_aov": 38.35, "email_aov": 31.05,
        "total_conversion": 0.0409, "web_conversion": 0.0002, "email_conversion": 0.0002,
        "notes": "Strong month — web impressions + sales scale up sharply.",
    }),
    (2026, "Apr", {
        "site": 1_647_453.27, "web": 79_915.41, "email": 37_248.79,
        "web_impressions": 8_633_109, "email_impressions": 7_834_380,
        "total_aov": 35.33, "web_aov": 40.32, "email_aov": 34.33,
        "total_conversion": 0.0354, "web_conversion": 0.0002, "email_conversion": 0.0001,
        "notes": "Web peak month — PDP rec slots likely key driver.",
    }),
    (2026, "May", {
        "site": 1_482_893.59, "web": 67_552.47, "email": 39_646.51,
        "web_impressions": 9_986_609, "email_impressions": 8_156_246,
        "total_aov": 33.91, "web_aov": 37.26, "email_aov": 34.12,
        "total_conversion": 0.0345, "web_conversion": 0.0002, "email_conversion": 0.0001,
        "notes": "Site sales down vs Apr; FR share held ~7%.",
    }),
    (2026, "Jun", {
        "site": 146_310.98, "web": 5_788.30, "email": 2_473.62,
        "web_impressions": 1_001_962, "email_impressions": 1_114_215,
        "total_aov": 32.39, "web_aov": 30.31, "email_aov": 31.71,
        "total_conversion": 0.0370, "web_conversion": 0.0002, "email_conversion": 0.0001,
        "notes": "MTD only (1–5 Jun) — update when month closes.",
    }),
    (2026, "Jul", None),
    (2026, "Aug", None),
    (2026, "Sep", None),
    (2026, "Oct", None),
    (2026, "Nov", None),
    (2026, "Dec", None),
]

COLUMNS = [
    ("Year", 8, "period"),
    ("Month", 10, "period"),
    ("Total Site Sales (£)", 18, "overview"),
    ("Total FR Attributed Sales (£)", 22, "overview"),
    ("FR Share of Site Sales (%)", 20, "overview"),
    ("FR Sales MoM (%)", 16, "overview"),
    ("Total AOV (£)", 14, "overview"),
    ("Total Conversion (%)", 18, "overview"),
    ("Web Impressions", 16, "web"),
    ("Web CTR (%)", 14, "web"),
    ("Web Content Conversion (%)", 22, "web"),
    ("Web Attributed Sales (£)", 22, "web"),
    ("Web Share of Site Sales (%)", 22, "web"),
    ("Web Content AOV (£)", 18, "web"),
    ("Email Widget Impressions", 22, "email"),
    ("Email Widget CTR (%)", 16, "email"),
    ("Email Content Conversion (%)", 26, "email"),
    ("Email Widget Attributed Sales (£)", 26, "email"),
    ("Email Share of Site Sales (%)", 24, "email"),
    ("Email Content AOV (£)", 18, "email"),
    ("Meeting Notes", 36, "notes"),
]

GROUP_FILLS = {
    "period": "FFD9D9D9",
    "overview": "FFC6EFCE",
    "web": "FFBDD7EE",
    "email": "FFFFEB9C",
    "notes": "FFF2F2F2",
}

GROUP_LABELS = {
    "period": "Period",
    "overview": "Overview",
    "web": "Web",
    "email": "Email",
    "notes": "Commentary",
}

DATA_FIELDS = {
    "site": "Total Site Sales (£)",
    "web": "Web Attributed Sales (£)",
    "email": "Email Widget Attributed Sales (£)",
    "web_impressions": "Web Impressions",
    "email_impressions": "Email Widget Impressions",
    "total_aov": "Total AOV (£)",
    "web_aov": "Web Content AOV (£)",
    "email_aov": "Email Content AOV (£)",
    "total_conversion": "Total Conversion (%)",
    "web_conversion": "Web Content Conversion (%)",
    "email_conversion": "Email Content Conversion (%)",
    "notes": "Meeting Notes",
}

REQUIRED_METRICS = [k for k in DATA_FIELDS if k != "notes"]

CTR_DEPENDS = {
    "Web CTR (%)": ("web_impressions", "web_conversion"),
    "Email Widget CTR (%)": ("email_impressions", "email_conversion"),
}

HEATMAP_COLUMNS = [
    "Total Site Sales (£)",
    "Total FR Attributed Sales (£)",
    "FR Share of Site Sales (%)",
    "Total AOV (£)",
    "Total Conversion (%)",
    "Web Impressions",
    "Web CTR (%)",
    "Web Content Conversion (%)",
    "Web Attributed Sales (£)",
    "Web Share of Site Sales (%)",
    "Web Content AOV (£)",
    "Email Widget Impressions",
    "Email Widget CTR (%)",
    "Email Content Conversion (%)",
    "Email Widget Attributed Sales (£)",
    "Email Share of Site Sales (%)",
    "Email Content AOV (£)",
]


def col(name: str) -> int:
    for idx, (header, _, _) in enumerate(COLUMNS, start=1):
        if header == name:
            return idx
    raise KeyError(name)


def letter(name: str) -> str:
    return get_column_letter(col(name))


def style_header_cell(cell, fill_hex: str):
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FF999999")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def column_groups():
    groups = []
    current = None
    start = 1
    for idx, (_, _, group) in enumerate(COLUMNS, start=1):
        if group != current:
            if current is not None:
                groups.append((current, start, idx - 1))
            current = group
            start = idx
    groups.append((current, start, len(COLUMNS)))
    return groups


def style_section_data_row(ws, row_idx: int, bold: bool = False):
    """Data rows — section background matches header, black text."""
    font = DATA_FONT_BOLD if bold else DATA_FONT
    for group, start, end in column_groups():
        fill = PatternFill("solid", fgColor=GROUP_FILLS[group])
        for c in range(start, end + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.font = font


def is_future_placeholder(year: int, data) -> bool:
    """2026 Jul–Dec rows — awaiting month close, no styling."""
    return year == 2026 and data is None


def highlight_missing(ws, row_idx: int, year: int, data):
    """Red = metric not available from FR screenshots (not legitimate zeros)."""
    if is_future_placeholder(year, data):
        return

    red = PatternFill("solid", fgColor=MISSING_FILL)
    metric_headers = {DATA_FIELDS[f] for f in REQUIRED_METRICS}

    if data is None:
        for header in metric_headers:
            ws.cell(row=row_idx, column=col(header)).fill = red
        for header in CTR_DEPENDS:
            ws.cell(row=row_idx, column=col(header)).fill = red
        for header in (
            "Total FR Attributed Sales (£)", "FR Share of Site Sales (%)",
            "Web Share of Site Sales (%)", "Email Share of Site Sales (%)",
        ):
            ws.cell(row=row_idx, column=col(header)).fill = red
        ws.cell(row=row_idx, column=col("Meeting Notes")).value = "Awaiting FR dashboard data."
        return

    provided = set(data.keys())
    for field in REQUIRED_METRICS:
        if field not in provided:
            ws.cell(row=row_idx, column=col(DATA_FIELDS[field])).fill = red

    for header, (imp_f, conv_f) in CTR_DEPENDS.items():
        if imp_f not in provided or conv_f not in provided:
            ws.cell(row=row_idx, column=col(header)).fill = red


def mom_formula(fr_col: str, row: int, prev_row: int) -> str:
    """Safe MoM — avoids #VALUE! when prior month has no FR sales."""
    return (
        f'=IFERROR(IF(AND(ISNUMBER({fr_col}{prev_row}),{fr_col}{prev_row}>0,'
        f'ISNUMBER({fr_col}{row}),{fr_col}{row}>0),'
        f'({fr_col}{row}-{fr_col}{prev_row})/{fr_col}{prev_row},""),"")'
    )


def apply_heatmap(ws, data_rows: list[int]):
    """MoM heatmap — only rows that have FR data (skips placeholders)."""
    for header in HEATMAP_COLUMNS:
        col_l = letter(header)
        for i in range(1, len(data_rows)):
            row = data_rows[i]
            prev = data_rows[i - 1]
            cell_ref = f"{col_l}{row}"
            ws.conditional_formatting.add(
                cell_ref,
                FormulaRule(
                    formula=[f'AND(ISNUMBER({col_l}{row}),ISNUMBER({col_l}{prev}),{col_l}{row}>{col_l}{prev})'],
                    fill=PatternFill("solid", fgColor=HEAT_UP_FILL),
                ),
            )
            ws.conditional_formatting.add(
                cell_ref,
                FormulaRule(
                    formula=[f'AND(ISNUMBER({col_l}{row}),ISNUMBER({col_l}{prev}),{col_l}{row}<{col_l}{prev})'],
                    fill=PatternFill("solid", fgColor=HEAT_DOWN_FILL),
                ),
            )

    mom_col = letter("FR Sales MoM (%)")
    for i in range(1, len(data_rows)):
        row = data_rows[i]
        cell_ref = f"{mom_col}{row}"
        ws.conditional_formatting.add(
            cell_ref,
            FormulaRule(
                formula=[f'AND(ISNUMBER({mom_col}{row}),{mom_col}{row}>0)'],
                fill=PatternFill("solid", fgColor=HEAT_UP_FILL),
            ),
        )
        ws.conditional_formatting.add(
            cell_ref,
            FormulaRule(
                formula=[f'AND(ISNUMBER({mom_col}{row}),{mom_col}{row}<0)'],
                fill=PatternFill("solid", fgColor=HEAT_DOWN_FILL),
            ),
        )


def build_data_sheet(ws):
    ws.title = "Monthly Performance"
    num_cols = len(COLUMNS)
    data_start = 3
    data_end = data_start + len(REPORT_ROWS) - 1
    total_row = data_end + 1

    # Totals row — all populated months (Jun 2025 – Jun 2026 MTD)
    period_rows = [
        data_start + i for i, (_, _, data) in enumerate(REPORT_ROWS) if data is not None
    ]
    period_ds = min(period_rows)
    period_de = max(period_rows)

    col_idx = 1
    group_ranges = []
    current_group = None
    start_col = 1

    for _, _, group in COLUMNS:
        if group != current_group:
            if current_group is not None:
                group_ranges.append((current_group, start_col, col_idx - 1))
            current_group = group
            start_col = col_idx
        col_idx += 1
    group_ranges.append((current_group, start_col, col_idx - 1))

    for group, start, end in group_ranges:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=GROUP_LABELS[group])
        style_header_cell(cell, GROUP_FILLS[group])
        for c in range(start + 1, end + 1):
            style_header_cell(ws.cell(row=1, column=c), GROUP_FILLS[group])

    for idx, (header, width, group) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        style_header_cell(cell, GROUP_FILLS[group])
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "C3"

    def C(name: str) -> int:
        return col(name)

    def L(name: str) -> str:
        return letter(name)

    c_site = L("Total Site Sales (£)")
    c_fr = L("Total FR Attributed Sales (£)")
    c_web_imp = L("Web Impressions")
    c_web_ctr = L("Web CTR (%)")
    c_web_conv = L("Web Content Conversion (%)")
    c_web_sales = L("Web Attributed Sales (£)")
    c_email_imp = L("Email Widget Impressions")
    c_email_ctr = L("Email Widget CTR (%)")
    c_email_conv = L("Email Content Conversion (%)")
    c_email_sales = L("Email Widget Attributed Sales (£)")
    c_total_aov = L("Total AOV (£)")
    c_total_conv = L("Total Conversion (%)")
    c_web_aov = L("Web Content AOV (£)")
    c_email_aov = L("Email Content AOV (£)")

    for i, (year, month, data) in enumerate(REPORT_ROWS):
        row_idx = data_start + i
        label = f"{month} (MTD)" if year == MTD_YEAR and month == MTD_MONTH else month
        ws.cell(row=row_idx, column=1, value=year)
        ws.cell(row=row_idx, column=2, value=label)

        if data:
            for field, header in DATA_FIELDS.items():
                if field in data:
                    ws.cell(row=row_idx, column=C(header), value=data[field])

        r = row_idx
        prev = r - 1
        ws.cell(row=r, column=C("Total FR Attributed Sales (£)"),
                value=f"={c_web_sales}{r}+{c_email_sales}{r}")
        ws.cell(row=r, column=C("FR Share of Site Sales (%)"),
                value=f'=IF({c_site}{r}>0,{c_fr}{r}/{c_site}{r},"")')
        if i == 0:
            ws.cell(row=r, column=C("FR Sales MoM (%)"), value="")  # first month — no prior period
        else:
            ws.cell(row=r, column=C("FR Sales MoM (%)"), value=mom_formula(c_fr, r, prev))
        ws.cell(row=r, column=C("Web Share of Site Sales (%)"),
                value=f'=IF({c_site}{r}>0,{c_web_sales}{r}/{c_site}{r},"")')
        ws.cell(row=r, column=C("Email Share of Site Sales (%)"),
                value=f'=IF({c_site}{r}>0,{c_email_sales}{r}/{c_site}{r},"")')
        ws.cell(row=r, column=C("Web CTR (%)"),
                value=f'=IF({c_web_imp}{r}>0,{c_web_conv}{r}/{CLICK_TO_ORDER},"")')
        ws.cell(row=r, column=C("Email Widget CTR (%)"),
                value=f'=IF({c_email_imp}{r}>0,{c_email_conv}{r}/{CLICK_TO_ORDER},0)')

    # Period totals row (Jun 2025 – Jun 2026 MTD — matches ~£9.5m site sales)
    tr = total_row
    ds, de = period_ds, period_de
    ws.cell(row=tr, column=1, value="")
    ws.cell(row=tr, column=2, value="Period Total")

    sum_cols = [
        "Total Site Sales (£)", "Web Impressions", "Web Attributed Sales (£)",
        "Email Widget Impressions", "Email Widget Attributed Sales (£)",
    ]
    for header in sum_cols:
        col_l = L(header)
        ws.cell(row=tr, column=col(header), value=f"=SUM({col_l}{ds}:{col_l}{de})")

    ws.cell(row=tr, column=C("Total FR Attributed Sales (£)"),
            value=f"=SUM({c_fr}{ds}:{c_fr}{de})")
    ws.cell(row=tr, column=C("FR Sales MoM (%)"),
            value=f'=IFERROR(IF(AND(ISNUMBER({c_fr}{ds}),{c_fr}{ds}>0,ISNUMBER({c_fr}{de}),{c_fr}{de}>0),({c_fr}{de}-{c_fr}{ds})/{c_fr}{ds},""),"")')
    ws.cell(row=tr, column=C("FR Share of Site Sales (%)"),
            value=f'=IF({c_site}{tr}>0,{c_fr}{tr}/{c_site}{tr},"")')
    ws.cell(row=tr, column=C("Web Share of Site Sales (%)"),
            value=f'=IF({c_site}{tr}>0,{c_web_sales}{tr}/{c_site}{tr},"")')
    ws.cell(row=tr, column=C("Email Share of Site Sales (%)"),
            value=f'=IF({c_site}{tr}>0,{c_email_sales}{tr}/{c_site}{tr},"")')
    ws.cell(row=tr, column=C("Meeting Notes"),
            value="Jun 2025 – Jun 2026 MTD. F = FR growth Jun'25→Jun'26. Attribution: 24h.")

    ws.cell(row=tr, column=C("Total AOV (£)"),
            value=f'=IF(SUM({c_site}{ds}:{c_site}{de})>0,SUMPRODUCT({c_site}{ds}:{c_site}{de},{c_total_aov}{ds}:{c_total_aov}{de})/SUM({c_site}{ds}:{c_site}{de}),"")')
    ws.cell(row=tr, column=C("Total Conversion (%)"),
            value=f'=IF(SUM({c_site}{ds}:{c_site}{de})>0,SUMPRODUCT({c_site}{ds}:{c_site}{de},{c_total_conv}{ds}:{c_total_conv}{de})/SUM({c_site}{ds}:{c_site}{de}),"")')
    ws.cell(row=tr, column=C("Web Content AOV (£)"),
            value=f'=IF(SUM({c_web_sales}{ds}:{c_web_sales}{de})>0,SUMPRODUCT({c_web_sales}{ds}:{c_web_sales}{de},{c_web_aov}{ds}:{c_web_aov}{de})/SUM({c_web_sales}{ds}:{c_web_sales}{de}),"")')
    ws.cell(row=tr, column=C("Email Content AOV (£)"),
            value=f'=IF(SUM({c_email_sales}{ds}:{c_email_sales}{de})>0,SUMPRODUCT({c_email_sales}{ds}:{c_email_sales}{de},{c_email_aov}{ds}:{c_email_aov}{de})/SUM({c_email_sales}{ds}:{c_email_sales}{de}),"")')
    ws.cell(row=tr, column=C("Web Content Conversion (%)"),
            value=f'=IF(SUM({c_web_imp}{ds}:{c_web_imp}{de})>0,SUMPRODUCT({c_web_imp}{ds}:{c_web_imp}{de},{c_web_conv}{ds}:{c_web_conv}{de})/SUM({c_web_imp}{ds}:{c_web_imp}{de}),"")')
    ws.cell(row=tr, column=C("Email Content Conversion (%)"),
            value=f'=IF(SUM({c_email_imp}{ds}:{c_email_imp}{de})>0,SUMPRODUCT({c_email_imp}{ds}:{c_email_imp}{de},{c_email_conv}{ds}:{c_email_conv}{de})/SUM({c_email_imp}{ds}:{c_email_imp}{de}),"")')
    ws.cell(row=tr, column=C("Web CTR (%)"),
            value=f'=IF(SUM({c_web_imp}{ds}:{c_web_imp}{de})>0,SUMPRODUCT({c_web_imp}{ds}:{c_web_imp}{de},{c_web_conv}{ds}:{c_web_conv}{de})/SUM({c_web_imp}{ds}:{c_web_imp}{de})/{CLICK_TO_ORDER},"")')
    ws.cell(row=tr, column=C("Email Widget CTR (%)"),
            value=f'=IF(SUM({c_email_imp}{ds}:{c_email_imp}{de})>0,SUMPRODUCT({c_email_imp}{ds}:{c_email_imp}{de},{c_email_conv}{ds}:{c_email_conv}{de})/SUM({c_email_imp}{ds}:{c_email_imp}{de})/{CLICK_TO_ORDER},"")')

    money_cols = {
        C("Total Site Sales (£)"), C("Total FR Attributed Sales (£)"),
        C("Web Attributed Sales (£)"), C("Email Widget Attributed Sales (£)"),
    }
    aov_cols = {C("Total AOV (£)"), C("Web Content AOV (£)"), C("Email Content AOV (£)")}
    count_cols = {C("Web Impressions"), C("Email Widget Impressions")}
    pct_cols = {
        C("FR Share of Site Sales (%)"), C("FR Sales MoM (%)"), C("Total Conversion (%)"),
        C("Web CTR (%)"), C("Web Content Conversion (%)"), C("Web Share of Site Sales (%)"),
        C("Email Widget CTR (%)"), C("Email Content Conversion (%)"),
        C("Email Share of Site Sales (%)"),
    }

    for row in range(data_start, total_row + 1):
        for c in money_cols:
            ws.cell(row=row, column=c).number_format = "£#,##0"
        for c in aov_cols:
            ws.cell(row=row, column=c).number_format = "£#,##0.00"
        for c in count_cols:
            ws.cell(row=row, column=c).number_format = "#,##0"
        for c in pct_cols:
            ws.cell(row=row, column=c).number_format = "0.00%"

    for row in range(data_start, total_row + 1):
        ws.cell(row=row, column=C("Meeting Notes")).alignment = Alignment(wrap_text=True, vertical="top")

    # Section backgrounds on all data + YTD rows (matches header colours)
    for row in range(data_start, total_row + 1):
        style_section_data_row(ws, row, bold=(row == total_row))

    for i, (year, month, data) in enumerate(REPORT_ROWS):
        row_idx = data_start + i
        highlight_missing(ws, row_idx, year, data)
        if year == MTD_YEAR and month == MTD_MONTH:
            for c in range(1, 3):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = PatternFill("solid", fgColor=MTD_FILL)
                cell.font = DATA_FONT

    data_rows = [
        data_start + i for i, (_, _, d) in enumerate(REPORT_ROWS) if d is not None
    ]
    apply_heatmap(ws, data_rows)


def _slot_num(value) -> float:
    if value in (None, "", "N/A", "n/a"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _slot_brand(name: str, slot_id: str) -> str:
    n, i = (name or "").lower(), (slot_id or "").lower()
    if "hrk" in i or "harkness" in n:
        return "Roses (Harkness)"
    if i.startswith("ms-") or n.startswith("ms ") or n.startswith("ms-"):
        return "Mailshop"
    if (
        i.startswith("yg-")
        or i.startswith("yougarden")
        or i.startswith("sale-strip")
        or i.startswith("0001-yg")
        or n.startswith("yg ")
        or "yougarden" in n
        or " - yg - " in n
        or n == "sale strip banner"
    ):
        return "You Garden"
    if i.startswith("gd-") or n.startswith("gd ") or "gardening direct" in n:
        return "Gardening Direct"
    return "Other"


def _slot_group_name(name: str) -> str:
    return re.sub(r"\s+EM[A-Z0-9]+$", "", (name or "").strip(), flags=re.I)


def load_web_slot_summary():
    """Aggregate FR slot export — web rows only (email types excluded)."""
    if not SLOTS_CSV.exists():
        raise FileNotFoundError(f"Slot CSV not found: {SLOTS_CSV}")

    by_brand_slot = defaultdict(float)
    with SLOTS_CSV.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            if row.get("Type") != "Web":
                continue
            revenue = _slot_num(row.get("On-site Revenue")) + _slot_num(
                row.get("Off-site Revenue")
            )
            if revenue <= 0:
                continue
            brand = _slot_brand(row.get("Name", ""), row.get("ID", ""))
            slot = _slot_group_name(row.get("Name", ""))
            by_brand_slot[(brand, slot)] += revenue

    brand_totals = defaultdict(float)
    for (brand, _), revenue in by_brand_slot.items():
        brand_totals[brand] += revenue

    grand_total = sum(brand_totals.values())
    brand_order = sorted(brand_totals, key=lambda b: brand_totals[b], reverse=True)

    detail_rows = []
    for brand in brand_order:
        slots = [
            (slot, revenue)
            for (b, slot), revenue in by_brand_slot.items()
            if b == brand
        ]
        slots.sort(key=lambda item: item[1], reverse=True)
        for slot, revenue in slots:
            detail_rows.append({
                "brand": brand,
                "slot": slot,
                "revenue": revenue,
                "pct_brand": revenue / brand_totals[brand] if brand_totals[brand] else 0,
                "pct_total": revenue / grand_total if grand_total else 0,
            })

    brand_summary = [
        {
            "brand": brand,
            "revenue": brand_totals[brand],
            "pct_total": brand_totals[brand] / grand_total if grand_total else 0,
        }
        for brand in brand_order
    ]
    return grand_total, brand_summary, detail_rows


def build_slots_sheet(ws):
    ws.title = "Web Slots"
    grand_total, brand_summary, detail_rows = load_web_slot_summary()

    web_fill = GROUP_FILLS["web"]
    thin = Side(style="thin", color="FF999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    title = ws.cell(row=1, column=1, value=f"Web Slot Performance — {SLOTS_PERIOD}")
    title.font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    subtitle = ws.cell(
        row=2,
        column=1,
        value="Web slots only. Email campaign slots excluded. Revenue = on-site + off-site.",
    )
    subtitle.font = Font(size=10, italic=True)
    ws.merge_cells("A2:E2")

    summary_headers = ["Brand", "Web Attributed Revenue (£)", "% of Total Web"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        style_header_cell(cell, web_fill)

    summary_start = 5
    for i, item in enumerate(brand_summary):
        row = summary_start + i
        ws.cell(row=row, column=1, value=item["brand"])
        ws.cell(row=row, column=2, value=item["revenue"])
        ws.cell(row=row, column=3, value=item["pct_total"])
        for c in range(1, 4):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=web_fill)
            cell.font = DATA_FONT
            cell.border = border

    total_row = summary_start + len(brand_summary)
    ws.cell(row=total_row, column=1, value="Total Web")
    ws.cell(row=total_row, column=2, value=grand_total)
    ws.cell(row=total_row, column=3, value=1 if grand_total else 0)
    for c in range(1, 4):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = PatternFill("solid", fgColor=web_fill)
        cell.font = DATA_FONT_BOLD
        cell.border = border

    detail_header_row = total_row + 2
    detail_headers = [
        "Brand",
        "Slot",
        "Attributed Revenue (£)",
        "% of Brand",
        "% of Total Web",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = ws.cell(row=detail_header_row, column=col_idx, value=header)
        style_header_cell(cell, web_fill)

    detail_start = detail_header_row + 1
    for i, item in enumerate(detail_rows):
        row = detail_start + i
        ws.cell(row=row, column=1, value=item["brand"])
        ws.cell(row=row, column=2, value=item["slot"])
        ws.cell(row=row, column=3, value=item["revenue"])
        ws.cell(row=row, column=4, value=item["pct_brand"])
        ws.cell(row=row, column=5, value=item["pct_total"])
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=web_fill)
            cell.font = DATA_FONT
            cell.border = border
            if c == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in range(summary_start, total_row + 1):
        ws.cell(row=row, column=2).number_format = "£#,##0"
        ws.cell(row=row, column=3).number_format = "0.0%"

    for row in range(detail_start, detail_start + len(detail_rows)):
        ws.cell(row=row, column=3).number_format = "£#,##0"
        ws.cell(row=row, column=4).number_format = "0.0%"
        ws.cell(row=row, column=5).number_format = "0.0%"

    ws.freeze_panes = "A5"


def build_readme_sheet(ws):
    ws.title = "How to use"
    ws.column_dimensions["A"].width = 100

    lines = [
        "Fresh Relevance — Monthly Performance Report (You Garden)",
        "",
        "Layout: 2025 Jun–Dec, then 2026 Jan–Dec. Period Total = all populated months.",
        "",
        "Attribution: Default (Product Only), 24-hour window.",
        "",
        "FR Sales MoM (%)",
        "• Blank when prior month has no FR sales (avoids #VALUE! errors)",
        "• Jun 2025 and Jan 2026 compare to prior month when data exists",
        "",
        "Formatting",
        "• Faded green = up vs prior month | Faded red = down (heatmap)",
        "• Pink/red cells = missing data (no FR screenshot available)",
        "• 2026 Jul–Dec = plain rows (no colour) until data available",
        "• Cell borders in Excel = selection outline, not part of the file",
        "• Section colours on all rows match headers (grey/green/blue/yellow)",
        "• Jun 2026 (MTD) = yellow period columns only",
        "• Period Total = bold, sums Jun 2025 – Jun 2026 (not 2026 calendar YTD only)",
        "",
        "Web Slots tab",
        f"• Period: {SLOTS_PERIOD} — from FR slot-level export",
        "• Web slots only — email campaign slots are excluded",
        "• Slot names grouped (EM#### email suffixes stripped where present)",
        "• Revenue = on-site + off-site attributed revenue per slot",
        f"• Source CSV: {SLOTS_CSV.name} — replace and re-run script to refresh",
    ]

    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)


def main():
    wb = Workbook()
    build_data_sheet(wb.active)
    build_slots_sheet(wb.create_sheet())
    build_readme_sheet(wb.create_sheet())
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
